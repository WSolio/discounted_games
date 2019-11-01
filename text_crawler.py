from selenium import webdriver
from bs4 import BeautifulSoup

import time
import datetime

from openpyxl import load_workbook


options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")

options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")





driver = webdriver.Chrome('C:/chromedriver')
driver.get('https://play.google.com/store/apps/collection/cluster?clp=ChwKGgoUdG9wc2VsbGluZ19wYWlkX0dBTUUQBxgD:S:ANO1ljI3NIA&gsr=Ch4KHAoaChR0b3BzZWxsaW5nX3BhaWRfR0FNRRAHGAM%3D:S:ANO1ljKgPH8&gsr=Ch4KHAoaChR0b3BzZWxsaW5nX3BhaWRfR0FNRRAHGAM%3D:S:ANO1ljKgPH8')


for i in range(4):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(4)

game_titles = driver.find_elements_by_class_name('WsMG1c') ###게임 제목
game_publishers = driver.find_elements_by_class_name('KoLSrc') ###게임 제작사
game_origin_prices = driver.find_elements_by_class_name('SUZt4c') ###게임 원래 가격 <---이 값이 존재하는 게임은 할인중인 게임
game_saling_prices = driver.find_elements_by_class_name('VfPpfd') ###게임 세일 가격 <---할인되거나, 일반적인 가격
game_prices = driver.find_elements_by_class_name('LCATme') ###게임 종합 가격 <---둘 다 포함되는 영역
game_links = driver.find_elements_by_class_name('poRVub')

#for game_link in game_links:
#    print(game_link.get_attribute("href"))

a = []

b = []
b1 = []

c = []
d = []
e = []
f = []
g = []

for game_title in game_titles:
    title = game_title.text
    a.append(title)

for game_publisher in game_publishers:
    publisher = game_publisher.text
    b1.append(publisher)
for i in range(len(b1)):
    if i % 2 ==1:
        b.append(b1[i])
        i = i+1

for game_saling_price in game_saling_prices:
    price_saling = game_saling_price.text
    c.append(price_saling)

for game_price in game_prices:
    price = game_price.text
    d.append(price)

for game_origin_price in game_origin_prices:
    price_origin = game_origin_price.text
    e.append(price_origin)
for game_link in game_links:
    f.append(game_link.get_attribute("href"))


games1=[]
n1=0
for n in range(len(a)):
    games1.append([a[n]]+[b[n]]+[c[n]]+[d[n]]+[f[n]])
    n1=n1+1
games = []

n=0
m=0
for n in range(200):
    if '0₩' in games1[n][3]:
        games.append(games1[n] + [e[m]])
        m = m + 1


titles = []
publishers=[]
o_prices = []
s_prices =[]
g_links =[]

n=0
for n in range(len(games)):
    titles.append(games[n][0])
    publishers.append(games[n][1])
    o_prices.append(games[n][5])
    s_prices.append(games[n][2])
    g_links.append(games[n][4])
    n=n+1

i=0
#for i in range(len(titles)):
    #print(titles[i],'\n','                         -published by', publishers[i], '\n', '   ', o_prices[i], '에서', s_prices[i],'로 할인 중!', '\n')
#    i=i+1



work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']


for i in range(len(titles)):
    work_sheet.cell(row=1, column=1, value='날두하자')
    work_sheet.cell(row=1, column=2, value='제목')
    work_sheet.cell(row=1, column=3, value='제작사')
    work_sheet.cell(row=1, column=4, value='가격정보')
    work_sheet.cell(row=1, column=5, value='링크')
    work_sheet.cell(row=i+2, column=1, value=i+1)
    work_sheet.cell(row=i+2, column=2, value=titles[i])
    work_sheet.cell(row=i+2, column=3, value=publishers[i])
    work_sheet.cell(row=i+2, column=4, value= o_prices[i] + ' 에서 '+ s_prices[i] +' 로 할인 중!')
    work_sheet.cell(row=i+2, column=5, value=g_links[i])
    i = i +1

work_book.save('오늘의 게임 날두 리스트.xlsx')



driver.quit()